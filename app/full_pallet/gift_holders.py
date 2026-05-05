from __future__ import annotations

import io
import logging
import math
import re
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from PIL import Image

from app.shared.constants import DIGITS_RE, IMAGE_ANCHOR_ROW_0BASED
from app.shared.models import GiftHolder
from app.shared.text_utils import _coerce_int, _norm_header

logger = logging.getLogger(__name__)


def _coerce_item_no(v: object) -> Optional[str]:
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    s = re.sub(r"\.0$", "", str(v).strip())
    s = DIGITS_RE.sub("", s)
    return s.zfill(6) if s else None



def _img_anchor_col(img) -> Optional[int]:
    try:
        anchor = getattr(img, "anchor", None)
        if anchor is None:
            return None
        _from = getattr(anchor, "_from", None)
        if _from is not None:
            return int(_from.col)
        frm = getattr(anchor, "from_", None)
        if frm is not None:
            return int(frm.col)
    except Exception:
        return None
    return None


def _img_anchor_row(img) -> Optional[int]:
    try:
        anchor = getattr(img, "anchor", None)
        if anchor is None:
            return None
        _from = getattr(anchor, "_from", None)
        if _from is not None:
            return int(_from.row)
        frm = getattr(anchor, "from_", None)
        if frm is not None:
            return int(frm.row)
    except Exception:
        return None
    return None


def _img_pixel_size(img) -> Tuple[int, int]:
    try:
        w = int(float(getattr(img, "width", 0) or 0))
        h = int(float(getattr(img, "height", 0) or 0))
        return max(0, w), max(0, h)
    except Exception:
        return 0, 0



def _img_bytes_and_ext(img) -> Tuple[Optional[bytes], Optional[str]]:
    try:
        raw = None
        if hasattr(img, "_data"):
            raw = img._data()
        elif hasattr(img, "ref"):
            ref = img.ref
            if isinstance(ref, bytes):
                raw = ref
            elif hasattr(ref, "read"):
                raw = ref.read()
        if not raw:
            return None, None

        fmt = str(getattr(img, "format", "") or "").strip().lower()
        ext = fmt if fmt in {"png", "jpg", "jpeg", "gif", "bmp"} else "png"

        try:
            im = Image.open(io.BytesIO(raw)).convert("RGBA")
            out = io.BytesIO()
            im.save(out, format="PNG")
            return out.getvalue(), "png"
        except Exception:
            return raw, ext
    except Exception:
        return None, None



def _extract_ws_images_by_col(ws) -> Dict[int, List[dict]]:
    """
    Map 0-based Excel column index -> all embedded image candidates in that column.
    """
    out: Dict[int, List[dict]] = {}
    for img in getattr(ws, "_images", []) or []:
        col = _img_anchor_col(img)
        if col is None:
            continue
        row = _img_anchor_row(img)
        img_bytes, img_ext = _img_bytes_and_ext(img)
        if not img_bytes:
            continue
        width, height = _img_pixel_size(img)
        out.setdefault(col, []).append(
            {
                "col": int(col),
                "row": int(row) if row is not None else -1,
                "width": width,
                "height": height,
                "image_bytes": img_bytes,
                "image_ext": img_ext or "png",
            }
        )
    return out



def _nearest_image_for_col(
    images_by_col: Dict[int, List[dict]],
    target_col: int,
    preferred_row: int,
    top_band_min_row: int,
    top_band_max_row: int,
    max_distance: int = 3,
) -> Tuple[Optional[bytes], Optional[str]]:
    candidates: List[dict] = []
    for col, col_candidates in images_by_col.items():
        dist = abs(int(col) - int(target_col))
        if dist > max_distance:
            continue
        candidates.extend(col_candidates or [])
    if not candidates:
        return None, None

    def _score(c: dict) -> Tuple[int, int, int, int, int]:
        row = int(c.get("row", -1))
        col = int(c.get("col", target_col))
        w = int(c.get("width", 0))
        h = int(c.get("height", 0))
        area = max(0, w) * max(0, h)
        in_top_band = top_band_min_row <= row <= top_band_max_row
        tiny = area < 2500
        return (
            0 if in_top_band else 1,
            1 if tiny else 0,
            abs(col - int(target_col)),
            area if area > 0 else 10**9,
            abs(row - int(preferred_row)) if row >= 0 else 10**9,
        )

    best = min(candidates, key=_score)
    return best.get("image_bytes"), best.get("image_ext")



def _image_for_col_span(
    images_by_col: Dict[int, List[dict]],
    start_col: int,
    end_col: int,
    preferred_row: int,
    top_band_min_row: int,
    top_band_max_row: int,
) -> Tuple[Optional[bytes], Optional[str]]:
    in_span = [
        cand
        for col, payloads in images_by_col.items()
        if start_col <= int(col) <= end_col
        for cand in (payloads or [])
    ]
    if in_span:
        # Prefer a top-band candidate with realistic card size over large mockup/render images.
        def _score_in_span(c: dict) -> Tuple[int, int, int, int, int]:
            row = int(c.get("row", -1))
            col = int(c.get("col", start_col))
            w = int(c.get("width", 0))
            h = int(c.get("height", 0))
            area = max(0, w) * max(0, h)
            in_top_band = top_band_min_row <= row <= top_band_max_row
            tiny = area < 2500
            return (
                0 if in_top_band else 1,
                1 if tiny else 0,
                max(0, col - int(start_col)),
                area if area > 0 else 10**9,
                abs(row - int(preferred_row)) if row >= 0 else 10**9,
            )

        best = min(in_span, key=_score_in_span)
        return best.get("image_bytes"), best.get("image_ext")
    center_col = (start_col + end_col) // 2
    return _nearest_image_for_col(
        images_by_col,
        center_col,
        preferred_row=preferred_row,
        top_band_min_row=top_band_min_row,
        top_band_max_row=top_band_max_row,
        max_distance=2,
    )



def _first_nonempty_in_span(ws, row_idx: int, start_col: int, end_col: int) -> Optional[str]:
    if row_idx < 0:
        return None
    for c in range(start_col, end_col + 1):
        val = ws.cell(row=row_idx + 1, column=c + 1).value
        if val is None:
            continue
        text = str(val).strip()
        if text:
            return text
    return None



def _first_int_in_span(ws, row_idx: int, start_col: int, end_col: int) -> Optional[int]:
    if row_idx < 0:
        return None
    for c in range(start_col, end_col + 1):
        val = ws.cell(row=row_idx + 1, column=c + 1).value
        if val is None:
            continue
        text = str(val).strip()
        if not text:
            continue
        digits = re.sub(r"[^\d]", "", text)
        if digits:
            try:
                return int(digits)
            except Exception:
                pass
    return None



def _gift_holder_slot_debug(slot: dict) -> dict:
    return {
        "header": str(slot.get("header", "") or ""),
        "segment_index": int(slot.get("segment_index", -1)),
        "start_col": int(slot.get("start_col", -1)),
        "end_col": int(slot.get("end_col", -1)),
        "item_no": str(slot.get("item_no", "") or ""),
        "qty": slot.get("qty"),
        "name": str(slot.get("name", "") or ""),
        "reject_reason": str(slot.get("reject_reason", "") or ""),
    }



def _extract_top_holder_slots(
    ws,
    images_by_col: Dict[int, List[dict]],
    qty_row: int,
    item_row: int,
    desc_row: int,
) -> List[dict]:
    """
    Extract holder slots from the top FULL PALLET block:
    POCKET/PEG headers -> image -> qty/item/description below.
    """
    pocket_peg_row_1based = IMAGE_ANCHOR_ROW_0BASED
    marketing_row_1based = IMAGE_ANCHOR_ROW_0BASED + 1
    if pocket_peg_row_1based < 1:
        return []

    max_col = ws.max_column
    slot_starts: List[Tuple[int, str]] = []

    for cidx in range(max_col):
        pocket_raw = ws.cell(row=pocket_peg_row_1based, column=cidx + 1).value
        pocket_text = str(pocket_raw).strip() if pocket_raw is not None else ""
        pocket_text_u = pocket_text.upper()

        marketing_raw = ws.cell(row=marketing_row_1based, column=cidx + 1).value
        marketing_text = str(marketing_raw).strip() if marketing_raw is not None else ""
        marketing_text_u = re.sub(r"\s+", " ", marketing_text.upper().strip())

        if pocket_text_u.startswith("POCKET ") or pocket_text_u.startswith("PEG "):
            slot_starts.append((cidx, pocket_text))
        elif "MARKETING MESSAGE PANEL" in marketing_text_u:
            slot_starts.append((cidx, "__BREAK__"))

    if not slot_starts:
        return []

    slots: List[dict] = []
    debug_candidates: List[dict] = []
    real_headers: List[Tuple[int, str, int]] = []
    marketing_break_cols = [c for c, h in slot_starts if h == "__BREAK__"]
    segment_index = 0
    for c, h in slot_starts:
        if h == "__BREAK__":
            segment_index += 1
            continue
        real_headers.append((c, h, segment_index))

    for i, (start_col, header_text, seg_idx) in enumerate(real_headers):
        next_boundaries = [
            c for c, _ in slot_starts
            if c > start_col
        ]
        end_col = (min(next_boundaries) - 1) if next_boundaries else (max_col - 1)
        if end_col < start_col:
            end_col = start_col

        item_no = _first_nonempty_in_span(ws, item_row, start_col, end_col)
        item_no_digits = re.sub(r"[^\d]", "", str(item_no)) if item_no else ""
        qty = _first_int_in_span(ws, qty_row, start_col, end_col)
        name = _first_nonempty_in_span(ws, desc_row, start_col, end_col) or ""
        debug_candidate = {
            "header": header_text,
            "segment_index": seg_idx,
            "start_col": start_col,
            "end_col": end_col,
            "item_no": item_no_digits,
            "qty": qty,
            "name": name,
            "reject_reason": "" if item_no else "missing_item_no",
        }
        if marketing_break_cols and any(
            abs(start_col - break_col) <= 8 or abs(end_col - break_col) <= 8
            for break_col in marketing_break_cols
        ):
            debug_candidates.append(debug_candidate)

        if not item_no:
            continue

        # Prefer images anchored in the top holder band; this avoids selecting large
        # display/mockup renders that may share nearby columns on the sheet.
        preferred_row = max(0, marketing_row_1based - 1)
        top_band_min_row = max(0, preferred_row - 1)
        top_band_max_row = max(
            top_band_min_row + 2,
            qty_row if qty_row >= 0 else preferred_row + 8,
        )
        img_bytes, img_ext = _image_for_col_span(
            images_by_col,
            start_col,
            end_col,
            preferred_row=preferred_row,
            top_band_min_row=top_band_min_row,
            top_band_max_row=top_band_max_row,
        )

        slots.append(
            {
                "header": header_text,
                "slot_order": i,
                "segment_index": seg_idx,
                "start_col": start_col,
                "end_col": end_col,
                "item_no": item_no_digits,
                "qty": qty,
                "name": name,
                "image_bytes": img_bytes,
                "image_ext": img_ext,
            }
        )

    if marketing_break_cols or any(str(s.get("item_no", "")) == "109107" for s in slots):
        logger.warning(
            "full_pallet gift holder debug: marketing_break_cols=%s candidates_around_marketing=%s detected_109107=%s detected_109107_slots=%s",
            marketing_break_cols,
            [_gift_holder_slot_debug(c) for c in debug_candidates],
            any(str(s.get("item_no", "")) == "109107" for s in slots),
            [_gift_holder_slot_debug(s) for s in slots if str(s.get("item_no", "")) == "109107"],
        )

    return slots



def _partition_top_slots_into_sides(top_slots: List[dict]) -> Dict[str, List[dict]]:
    """
    Partition top holder slots into ordered side segments using
    MARKETING MESSAGE PANEL-delimited segment indices from extraction.
    """
    by_side: Dict[str, List[dict]] = {s: [] for s in "ABCD"}
    if not top_slots:
        return by_side

    slots_ordered = sorted(
        top_slots,
        key=lambda s: (
            int(s.get("segment_index", 0)),
            int(s.get("start_col", 0)),
            int(s.get("end_col", 0)),
        ),
    )
    segment_ids = sorted({int(s.get("segment_index", 0)) for s in slots_ordered})
    slots_by_segment: Dict[int, List[dict]] = {
        seg_id: [
            s for s in slots_ordered
            if int(s.get("segment_index", 0)) == seg_id
        ]
        for seg_id in segment_ids
    }

    expected_side_slots = 5
    seg_cursor = 0
    fallback_events: List[dict] = []
    side_debug_events: List[dict] = []

    for side in "ABCD":
        if seg_cursor >= len(segment_ids):
            break

        seg_id = segment_ids[seg_cursor]
        side_slots = list(slots_by_segment.get(seg_id, []))
        initial_count = len(side_slots)
        seg_cursor += 1

        added_item_ids: List[str] = []
        seen = {
            (
                str(s.get("item_no", "")).strip(),
                int(s.get("start_col", -1)),
                int(s.get("end_col", -1)),
            )
            for s in side_slots
        }

        while len(side_slots) < expected_side_slots and seg_cursor < len(segment_ids):
            next_seg_id = segment_ids[seg_cursor]
            next_slots = list(slots_by_segment.get(next_seg_id, []))
            considered = {
                "side": side,
                "initial_holder_count": initial_count,
                "current_holder_count": len(side_slots),
                "needed": expected_side_slots - len(side_slots),
                "candidate_segment": next_seg_id,
                "candidate_count": len(next_slots),
                "candidates": [_gift_holder_slot_debug(s) for s in next_slots],
                "fallback_used": False,
                "reject_reason": "",
            }
            if not next_slots:
                considered["reject_reason"] = "empty_candidate_segment"
                side_debug_events.append(considered)
                seg_cursor += 1
                continue

            needed = expected_side_slots - len(side_slots)
            considered["needed"] = needed

            # Edge case only: a marketing panel can split a small trailing run
            # from the side before it. Do not borrow part of a plausible next side.
            if len(next_slots) > needed:
                considered["reject_reason"] = "candidate_segment_larger_than_needed"
                side_debug_events.append(considered)
                break

            moved: List[dict] = []
            for slot in next_slots[:needed]:
                key = (
                    str(slot.get("item_no", "")).strip(),
                    int(slot.get("start_col", -1)),
                    int(slot.get("end_col", -1)),
                )
                if key in seen:
                    considered["reject_reason"] = "duplicate_candidate"
                    continue
                seen.add(key)
                moved.append(slot)
                added_item_ids.append(str(slot.get("item_no", "")).strip())

            side_slots.extend(moved)
            considered["fallback_used"] = bool(moved)
            considered["added_holder_ids"] = [str(s.get("item_no", "")).strip() for s in moved]
            considered["final_holder_count"] = len(side_slots)
            if not moved and not considered["reject_reason"]:
                considered["reject_reason"] = "no_unique_candidates"
            side_debug_events.append(considered)
            remaining = next_slots[needed:]
            if remaining:
                slots_by_segment[next_seg_id] = remaining
                break
            seg_cursor += 1

        by_side[side].extend(side_slots)
        if added_item_ids:
            fallback_events.append(
                {
                    "side": side,
                    "initial_holder_count": initial_count,
                    "fallback_used": True,
                    "added_holder_ids": added_item_ids,
                    "final_holder_count": len(side_slots),
                }
            )

    for side in "ABCD":
        by_side[side].sort(
            key=lambda s: (
                int(s.get("start_col", 0)),
                int(s.get("end_col", 0)),
            )
        )

    for event in fallback_events:
        logger.debug(
            "full_pallet gift holder marketing-panel fallback: side=%s initial_holder_count=%s fallback_used=%s added_holder_ids=%s final_holder_count=%s",
            event["side"],
            event["initial_holder_count"],
            event["fallback_used"],
            event["added_holder_ids"],
            event["final_holder_count"],
        )

    side_a_debug = [e for e in side_debug_events if e.get("side") == "A"]
    if side_a_debug or any(str(s.get("item_no", "")) == "109107" for s in slots_ordered):
        assigned_109107 = [
            {
                "side": side,
                "slot": _gift_holder_slot_debug(slot),
            }
            for side, side_slots in by_side.items()
            for slot in side_slots
            if str(slot.get("item_no", "")) == "109107"
        ]
        logger.warning(
            "full_pallet gift holder Side A fallback debug: initial_side_a_count=%s fallback_used=%s fallback_considered=%s assigned_109107=%s final_side_a_count=%s",
            side_a_debug[0].get("initial_holder_count") if side_a_debug else len(by_side.get("A", [])),
            any(bool(e.get("fallback_used")) for e in side_a_debug),
            side_a_debug,
            assigned_109107,
            len(by_side.get("A", [])),
        )

    return by_side



def load_gift_card_holders(gift_bytes: bytes) -> Dict[str, List[GiftHolder]]:
    """Parse POG workbook holder section and return ordered per-side holders."""

    try:
        xls = pd.read_excel(io.BytesIO(gift_bytes), sheet_name=None, header=None)
    except Exception as e:
        raise ValueError(f"Unable to read Gift Card Holders workbook: {e}")

    def find_header_row(df: pd.DataFrame, tokens: List[str], max_rows: int = 120) -> int:
        for i in range(min(len(df), max_rows)):
            row = df.iloc[i].astype(str).fillna("").str.upper().tolist()
            if all(any(tok in c for c in row) for tok in tokens):
                return i
        return -1

    def normalize_headers(raw_headers: List[object]) -> List[str]:
        headers: List[str] = []
        seen: Dict[str, int] = {}
        for v in raw_headers:
            base = _norm_header(v)
            n = seen.get(base, 0) + 1
            seen[base] = n
            headers.append(base if n == 1 else f"{base}_{n}")
        return headers

    def pick_col(headers: List[str], tokens: List[str]) -> Optional[str]:
        up = [h.upper() for h in headers]
        for tok in tokens:
            for i, h in enumerate(up):
                if tok in h:
                    return headers[i]
        return None

    def parse_side(value: object, current_side: str) -> str:
        text = str(value or "").upper()
        m = re.search(r"SIDE\s*([A-D])", text)
        if m:
            return m.group(1)
        m = re.fullmatch(r"\s*([A-D])\s*", text)
        if m:
            return m.group(1)
        return current_side

    full_pallet_sheet = next(
        (name for name in xls.keys() if "FULL" in name.upper() and "PALLET" in name.upper()),
        None,
    )
    if full_pallet_sheet is None:
        raise ValueError("FULL PALLET sheet/table missing in workbook.")

    try:
        wb = load_workbook(io.BytesIO(gift_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"Unable to open Gift Card Holders workbook images: {e}")

    full_pallet_ws = None
    for ws in wb.worksheets:
        title_u = str(ws.title or "").upper()
        if "FULL" in title_u and "PALLET" in title_u:
            full_pallet_ws = ws
            break

    if full_pallet_ws is None:
        raise ValueError("Could not find FULL PALLET sheet in workbook.")

    images_by_col = _extract_ws_images_by_col(full_pallet_ws)

    desc_map: Dict[str, str] = {}
    for _, sheet_df in xls.items():
        hdr = find_header_row(sheet_df, ["ITEM", "DESCRIPTION"])
        if hdr < 0:
            continue
        cols = normalize_headers(sheet_df.iloc[hdr].tolist())
        data = sheet_df.iloc[hdr + 1 :].copy()
        data.columns = cols
        item_lookup_col = pick_col(cols, ["ITEM", "ITEM_#", "SKU"])
        desc_lookup_col = pick_col(cols, ["DESCRIPTION", "DESC", "NAME"])
        if not item_lookup_col or not desc_lookup_col:
            continue
        for _, row in data.iterrows():
            item_no = _coerce_item_no(row.get(item_lookup_col))
            if not item_no:
                continue
            desc = str(row.get(desc_lookup_col, "") or "").strip()
            if desc:
                desc_map[item_no] = desc

    full_df_raw = xls[full_pallet_sheet].copy()

    item_row = -1
    qty_row = -1
    desc_row = -1
    max_rows = len(full_df_raw)

    # First pass: detect the real ITEM # row in the top holder block.
    for i in range(max_rows):
        row_vals = [str(v or "").strip().upper() for v in full_df_raw.iloc[i].tolist()]
        if any("ITEM #" in v or v == "ITEM" for v in row_vals):
            numeric_items = sum(1 for v in row_vals if _coerce_item_no(v))
            if numeric_items >= 4:
                item_row = i
                break

    # Second pass: robustly locate QTY relative to ITEM row.
    if item_row >= 0:
        above_candidates = []
        for r in range(max(0, item_row - 3), item_row + 1):
            row_vals = [str(v or "").strip().upper() for v in full_df_raw.iloc[r].tolist()]
            if any(v == "QTY" or v.startswith("QTY") for v in row_vals):
                above_candidates.append(r)
        if above_candidates:
            qty_row = max(above_candidates)

        if qty_row < 0:
            # Fallback scan: pick nearest QTY row at/around ITEM row.
            nearest_dist = None
            nearest_row = -1
            for r in range(max_rows):
                row_vals = [str(v or "").strip().upper() for v in full_df_raw.iloc[r].tolist()]
                if any(v == "QTY" or v.startswith("QTY") for v in row_vals):
                    dist = abs(r - item_row)
                    if nearest_dist is None or dist < nearest_dist:
                        nearest_dist = dist
                        nearest_row = r
            qty_row = nearest_row

        for r in range(item_row, min(max_rows, item_row + 8)):
            row_vals = [str(v or "").strip().upper() for v in full_df_raw.iloc[r].tolist()]
            if any("DESCRIPTION" in v for v in row_vals):
                desc_row = r
                break

    top_slots = _extract_top_holder_slots(
        full_pallet_ws,
        images_by_col=images_by_col,
        qty_row=qty_row,
        item_row=item_row,
        desc_row=desc_row,
    )
    logger.debug(
        "full_pallet top holder parse: item_row=%s qty_row=%s desc_row=%s top_slots=%s headers=%s items=%s",
        item_row,
        qty_row,
        desc_row,
        len(top_slots),
        [str(s.get("header", "")) for s in top_slots[:8]],
        [str(s.get("item_no", "")) for s in top_slots[:8]],
    )

    holders: Dict[str, List[GiftHolder]] = {s: [] for s in "ABCD"}
    if top_slots:
        slots_by_side = _partition_top_slots_into_sides(top_slots)
        for side in "ABCD":
            for slot in slots_by_side[side]:
                item_no = str(slot["item_no"]).strip()
                if not item_no:
                    continue
                holders[side].append(
                    GiftHolder(
                        side=side,
                        item_no=item_no,
                        name=str(slot["name"] or "").strip(),
                        qty=slot["qty"],
                        image_bytes=slot["image_bytes"],
                        image_ext=slot["image_ext"],
                        slot_label=str(slot["header"] or "").strip() or None,
                        slot_order=int(slot["slot_order"]),
                        slot_start_col=int(slot["start_col"]),
                        slot_end_col=int(slot["end_col"]),
                    )
                )
        return holders

    header_row = find_header_row(full_df_raw, ["ITEM", "QTY"])
    if header_row < 0:
        raise ValueError("FULL PALLET holder table missing ITEM/QTY columns.")

    headers = normalize_headers(full_df_raw.iloc[header_row].tolist())
    full_df = full_df_raw.iloc[header_row + 1 :].copy()
    full_df.columns = headers

    item_col = pick_col(headers, ["ITEM", "ITEM_#", "SKU"])
    qty_col = pick_col(headers, ["QTY", "QUANTITY"])
    side_col = pick_col(headers, ["SIDE"])
    if item_col is None or qty_col is None:
        raise ValueError("FULL PALLET holder table missing ITEM/QTY columns.")

    current_side = "A"
    for _, row in full_df.iterrows():
        side_source = row.get(side_col) if side_col else " "
        current_side = parse_side(side_source, current_side)

        item_no = _coerce_item_no(row.get(item_col))
        if not item_no:
            row_text = " ".join(str(v or "") for v in row.tolist())
            current_side = parse_side(row_text, current_side)
            continue

        qty = _coerce_int(row.get(qty_col))
        name = desc_map.get(item_no) or "(missing description)"
        side = current_side if current_side in "ABCD" else "A"
        holders.setdefault(side, []).append(
                GiftHolder(
                    side=side,
                    item_no=item_no,
                    name=name,
                    qty=qty,
                    image_bytes=None,
                    image_ext=None,
                )
            )

    if not any(holders.values()):
        raise ValueError("No holder Item # rows found in FULL PALLET table.")

    non_empty = [s for s in "ABCD" if holders.get(s)]
    if non_empty == ["A"]:
        base_list = holders["A"]
        for s in "BCD":
            holders[s] = [
                GiftHolder(
                    side=s,
                    item_no=h.item_no,
                    name=h.name,
                    qty=h.qty,
                    image_bytes=h.image_bytes,
                    image_ext=h.image_ext,
                    slot_label=h.slot_label,
                    slot_order=h.slot_order,
                    slot_start_col=h.slot_start_col,
                    slot_end_col=h.slot_end_col,
                )
                for h in base_list
            ]
    return holders
